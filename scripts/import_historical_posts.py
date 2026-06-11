from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import sys
import uuid
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parents[1]
POST_DB_DIR = ROOT_DIR / "Post DB"
DEFAULT_EXCEL_PATH = POST_DB_DIR / "chatgptricks_posts.xlsx"
DEFAULT_COVERS_DIR = POST_DB_DIR / "covers"
FLOP_LIKES_BASELINE = 850


def main() -> None:
    parser = argparse.ArgumentParser(description="Import historical Instagram posts from Post DB.")
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL_PATH)
    parser.add_argument("--covers-dir", type=Path, default=DEFAULT_COVERS_DIR)
    parser.add_argument("--start", type=int, default=1)
    parser.add_argument("--limit", type=int, default=5)
    parser.add_argument(
        "--include-types",
        default="",
        help="Comma-separated source Type prefixes to import, for example: image,carousel.",
    )
    parser.add_argument(
        "--exclude-types",
        default="",
        help="Comma-separated source Type prefixes to skip, for example: reel,video.",
    )
    parser.add_argument("--duration-seconds", type=int, default=2)
    parser.add_argument("--ocr", action=argparse.BooleanOptionalAction, default=False)
    parser.add_argument("--analyze", action=argparse.BooleanOptionalAction, default=False)
    parser.add_argument("--force-analyze", action="store_true")
    args = parser.parse_args()
    if args.ocr:
        raise SystemExit(
            "Import OCR is disabled. Run scripts/run_modal_ocr_batch.py after 100 completed Post DB posts are ready."
        )

    os.chdir(ROOT_DIR)
    _load_dotenv(ROOT_DIR / ".env")
    sys.path.insert(0, str(ROOT_DIR / "backend"))

    from app.config import ANALYSIS_DIR, DATA_DIR, VIDEO_DIR
    from app.db import connect, init_db, utc_now
    from app.remote_tribe import RemoteTribeUnavailable, analyze_video_remote, remote_tribe_status
    from app.video import VideoConversionError, create_static_video

    init_db()
    rows = _read_posts(args.excel)
    selected, skipped = _select_rows(
        rows,
        start=args.start,
        limit=args.limit,
        include_types=_parse_type_filter(args.include_types),
        exclude_types=_parse_type_filter(args.exclude_types),
    )
    if not selected:
        raise SystemExit(f"No matching rows found from {args.start:04d}.")
    if skipped:
        print(json.dumps({"skipped": skipped}, indent=2), file=sys.stderr)
    if len(selected) < args.limit:
        print(
            f"Only found {len(selected)} matching rows from {args.start:04d}; requested {args.limit}.",
            file=sys.stderr,
        )

    imported: list[dict[str, Any]] = []
    for row in selected:
        source_row_number = int(row["#"])
        shortcode = _clean_text(row.get("Shortcode"))
        source_ref = f"chatgptricks:{source_row_number:04d}"
        cover_path = _cover_path(row, args.covers_dir)
        image_path = _copy_cover(cover_path, DATA_DIR)
        title = _title_from_caption(row.get("Caption")) or f"{source_row_number:04d} - {shortcode}"
        post_type = _post_type(row.get("Type"))
        published_at = _published_at(row.get("Post Date UTC"))
        likes = _optional_int(row.get("Likes"))
        if likes is None:
            likes = FLOP_LIKES_BASELINE
        comments = _optional_int(row.get("Comments"))
        caption = _clean_text(row.get("Caption")) or None
        now = utc_now()

        with connect() as conn:
            existing = conn.execute(
                "SELECT * FROM posts WHERE source_ref = ?",
                (source_ref,),
            ).fetchone()
            hook_text = None
            if existing:
                conn.execute(
                    """
                    UPDATE posts
                    SET title = ?, caption = ?, published_at = ?, likes = ?,
                        post_type_label = ?, source_row_number = ?, shortcode = ?,
                        comments = ?, hook_text = COALESCE(?, hook_text),
                        image_path = ?, original_filename = ?, updated_at = ?
                    WHERE id = ?
                    """,
                    (
                        title,
                        caption,
                        published_at,
                        likes,
                        post_type,
                        source_row_number,
                        shortcode,
                        comments,
                        hook_text,
                        str(image_path),
                        cover_path.name,
                        now,
                        int(existing["id"]),
                    ),
                )
                post_id = int(existing["id"])
                action = "updated"
            else:
                cursor = conn.execute(
                    """
                    INSERT INTO posts (
                        section, title, caption, published_at, likes,
                        person_label, company_label, post_type_label,
                        comments, hook_text, source_ref, source_row_number, shortcode,
                        image_path, original_filename, status,
                        progress_percent, progress_message, created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        "historical",
                        title,
                        caption,
                        published_at,
                        likes,
                        None,
                        None,
                        post_type,
                        comments,
                        hook_text,
                        source_ref,
                        source_row_number,
                        shortcode,
                        str(image_path),
                        cover_path.name,
                        "queued",
                        0,
                        "Imported; analysis pending",
                        now,
                        now,
                    ),
                )
                post_id = int(cursor.lastrowid)
                action = "inserted"
            _save_metadata_option(conn, "post_type", post_type)

        imported.append(
            {
                "post_id": post_id,
                "source_ref": source_ref,
                "shortcode": shortcode,
                "likes": likes,
                "comments": comments,
                "cover": cover_path.name,
                "type": post_type,
                "ocr_text": hook_text,
                "action": action,
            }
        )

    if args.analyze:
        remote = remote_tribe_status()
        if not remote["configured"]:
            raise SystemExit("REMOTE_TRIBE_URL is not configured; import succeeded but analysis did not run.")
        for item in imported:
            post_id = int(item["post_id"])
            with connect() as conn:
                post = conn.execute("SELECT * FROM posts WHERE id = ?", (post_id,)).fetchone()
            if not post:
                continue
            already_done = post["status"] == "completed" and post["analysis_summary"]
            if already_done and not args.force_analyze:
                item["analysis"] = "skipped_completed"
                print(f"Skipped {item['source_ref']}; analysis already completed.", file=sys.stderr, flush=True)
                continue
            print(
                f"Analyzing {item['source_ref']} ({item['cover']}) on remote GPU...",
                file=sys.stderr,
                flush=True,
            )
            try:
                _set_progress(post_id, 10, "Preparing cover", "running")
                video_path = VIDEO_DIR / f"{post_id}-{uuid.uuid4().hex}.mp4"
                _set_progress(post_id, 22, "Converting image to video", "running")
                create_static_video(Path(post["image_path"]), video_path, duration_seconds=args.duration_seconds)
                _set_progress(post_id, 38, "Video ready; sending to remote GPU", "running")
                summary = analyze_video_remote(video_path, duration_seconds=args.duration_seconds)
                analysis_path = ANALYSIS_DIR / f"{post_id}-{uuid.uuid4().hex}.json"
                analysis_path.parent.mkdir(parents=True, exist_ok=True)
                analysis_path.write_text(json.dumps(summary, indent=2), encoding="utf-8")
                with connect() as conn:
                    conn.execute(
                        """
                        UPDATE posts
                        SET video_path = ?, analysis_path = ?, analysis_summary = ?,
                            status = ?, error = NULL, progress_percent = ?,
                            progress_message = ?, llm_report = NULL, updated_at = ?
                        WHERE id = ?
                        """,
                        (
                            str(video_path),
                            str(analysis_path),
                            json.dumps(summary),
                            "completed",
                            100,
                            "Complete",
                            utc_now(),
                            post_id,
                        ),
                )
                item["analysis"] = "completed"
                print(f"Completed {item['source_ref']}.", file=sys.stderr, flush=True)
            except (RemoteTribeUnavailable, VideoConversionError, Exception) as exc:
                with connect() as conn:
                    conn.execute(
                        """
                        UPDATE posts
                        SET status = ?, error = ?, progress_message = ?, updated_at = ?
                        WHERE id = ?
                        """,
                        ("failed", str(exc), "Analysis failed", utc_now(), post_id),
                    )
                item["analysis"] = "failed"
                item["error"] = str(exc)
                print(json.dumps({"imported": imported}, indent=2), file=sys.stderr)
                raise

    print(json.dumps({"imported": imported}, indent=2))


def _load_dotenv(path: Path) -> None:
    if not path.exists():
        return
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        os.environ.setdefault(key, value)


def _read_posts(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Posts"]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    records: list[dict[str, Any]] = []
    for values in rows:
        record = dict(zip(headers, values, strict=False))
        if _optional_int(record.get("#")) is None:
            continue
        records.append(record)
    return records


def _select_rows(
    rows: list[dict[str, Any]],
    *,
    start: int,
    limit: int,
    include_types: set[str],
    exclude_types: set[str],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    selected: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []
    candidates = sorted(
        (row for row in rows if int(row["#"]) >= start),
        key=lambda row: int(row["#"]),
    )
    for row in candidates:
        source_type = _type_key(row.get("Type"))
        reason = ""
        if include_types and source_type not in include_types:
            reason = "not_included_type"
        elif exclude_types and source_type in exclude_types:
            reason = "excluded_type"
        if reason:
            skipped.append(
                {
                    "row": int(row["#"]),
                    "type": _clean_text(row.get("Type")),
                    "reason": reason,
                }
            )
            continue
        selected.append(row)
        if len(selected) >= limit:
            break
    return selected, skipped


def _parse_type_filter(value: str) -> set[str]:
    return {
        _type_key(item)
        for item in value.split(",")
        if _type_key(item)
    }


def _type_key(value: Any) -> str:
    return _clean_text(value).split("(", 1)[0].strip().casefold()


def _cover_path(row: dict[str, Any], covers_dir: Path) -> Path:
    cover_file = _clean_text(row.get("Cover File"))
    if cover_file:
        candidate = covers_dir / Path(cover_file).name
        if candidate.exists():
            return candidate
    source_row_number = int(row["#"])
    matches = sorted(covers_dir.glob(f"{source_row_number:04d}_*"))
    if not matches:
        raise FileNotFoundError(f"Missing cover for row {source_row_number:04d} in {covers_dir}")
    return matches[0]


def _copy_cover(source: Path, data_dir: Path) -> Path:
    target_dir = data_dir / "uploads" / "imported-history"
    target_dir.mkdir(parents=True, exist_ok=True)
    target = target_dir / source.name
    if not target.exists() or source.stat().st_size != target.stat().st_size:
        shutil.copy2(source, target)
    return target


def _title_from_caption(value: Any) -> str:
    caption = _clean_text(value)
    if not caption:
        return ""
    for line in caption.splitlines():
        cleaned = " ".join(line.split())
        if cleaned:
            return cleaned[:120]
    return ""


def _post_type(value: Any) -> str | None:
    text = _clean_text(value)
    if not text:
        return None
    return text.split("(", 1)[0].strip()[:80] or None


def _published_at(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        timestamp = value.replace(tzinfo=UTC) if value.tzinfo is None else value.astimezone(UTC)
        return timestamp.isoformat(timespec="seconds")
    text = _clean_text(value)
    if not text:
        return None
    try:
        timestamp = datetime.fromisoformat(text).replace(tzinfo=UTC)
        return timestamp.isoformat(timespec="seconds")
    except ValueError:
        return text


def _optional_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return max(0, int(float(value)))
    except (TypeError, ValueError):
        return None


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _save_metadata_option(conn: Any, kind: str, label: str | None) -> None:
    if not label:
        return
    now = datetime.now(UTC).isoformat(timespec="seconds")
    conn.execute(
        """
        INSERT INTO metadata_options (kind, label, slug, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?)
        ON CONFLICT(kind, slug) DO UPDATE SET
            label = excluded.label,
            updated_at = excluded.updated_at
        """,
        (kind, label, _metadata_slug(label), now, now),
    )


def _metadata_slug(label: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", label.casefold()).strip("-")
    return slug or uuid.uuid4().hex


def _set_progress(post_id: int, percent: int, message: str, status: str) -> None:
    from app.db import connect, utc_now

    with connect() as conn:
        conn.execute(
            """
            UPDATE posts
            SET status = ?, error = NULL, progress_percent = ?,
                progress_message = ?, updated_at = ?
            WHERE id = ?
            """,
            (status, max(0, min(100, int(percent))), message, utc_now(), post_id),
        )


if __name__ == "__main__":
    main()
