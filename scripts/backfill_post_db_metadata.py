from __future__ import annotations

import argparse
import json
import os
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parents[1]
DEFAULT_EXCEL_PATH = ROOT_DIR / "Post DB" / "chatgptricks_posts.xlsx"


def main() -> None:
    parser = argparse.ArgumentParser(description="Backfill Post DB comments, captions, and OCR text.")
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL_PATH)
    parser.add_argument("--start", type=int, default=1)
    parser.add_argument("--limit", type=int, default=0, help="0 means no limit.")
    parser.add_argument("--ocr", action=argparse.BooleanOptionalAction, default=False)
    args = parser.parse_args()
    if args.ocr:
        raise SystemExit("OCR is Modal-only. Run scripts/run_modal_ocr_batch.py instead.")

    os.chdir(ROOT_DIR)
    _load_dotenv(ROOT_DIR / ".env")
    sys.path.insert(0, str(ROOT_DIR / "backend"))

    from app.db import connect, init_db, utc_now

    init_db()
    excel_rows = {
        int(row["#"]): row
        for row in _read_posts(args.excel)
        if _optional_int(row.get("#")) is not None
    }

    upper_bound = args.start + args.limit - 1 if args.limit else None
    with connect() as conn:
        post_rows = conn.execute(
            """
            SELECT id, source_row_number, image_path, hook_text
            FROM posts
            WHERE section = 'historical'
              AND source_row_number IS NOT NULL
              AND source_row_number >= ?
              AND (? IS NULL OR source_row_number <= ?)
            ORDER BY source_row_number
            """,
            (args.start, upper_bound, upper_bound),
        ).fetchall()

    pending: list[dict[str, Any]] = []
    for post in post_rows:
        row_number = int(post["source_row_number"])
        source = excel_rows.get(row_number)
        if not source:
            continue
        pending.append(
            {
                "id": int(post["id"]),
                "source_row_number": row_number,
                "image_path": post["image_path"],
                "comments": _optional_int(source.get("Comments")),
                "caption": _clean_text(source.get("Caption")) or None,
                "hook_text": None,
            }
        )

    updated: list[dict[str, Any]] = []
    with connect() as conn:
        for record in pending:
            conn.execute(
                """
                UPDATE posts
                SET comments = COALESCE(?, comments),
                    caption = COALESCE(?, caption),
                    hook_text = COALESCE(?, hook_text),
                    updated_at = ?
                WHERE id = ?
                """,
                (
                    record["comments"],
                    record["caption"],
                    record["hook_text"],
                    utc_now(),
                    record["id"],
                ),
            )
            updated.append(
                {
                    "id": record["id"],
                    "source_row_number": record["source_row_number"],
                    "comments": record["comments"],
                    "ocr_text": record["hook_text"],
                }
            )

    print(json.dumps({"updated": updated}, indent=2))


def _load_dotenv(path: Path) -> None:
    if not path.exists():
        return
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def _read_posts(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Posts"]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    records: list[dict[str, Any]] = []
    for values in rows:
        record = dict(zip(headers, values, strict=False))
        if _optional_int(record.get("#")) is None:
            continue
        records.append(record)
    return records


def _optional_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return max(0, int(float(value)))
    except (TypeError, ValueError):
        return None


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


if __name__ == "__main__":
    main()
