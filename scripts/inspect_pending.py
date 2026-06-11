import os
import sys
from pathlib import Path
from openpyxl import load_workbook

ROOT_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT_DIR / "backend"))

from app.db import connect

def _optional_int(value) -> int | None:
    if value is None or value == "":
        return None
    try:
        return max(0, int(float(value)))
    except (TypeError, ValueError):
        return None

def main():
    excel_path = ROOT_DIR / "Post DB" / "chatgptricks_posts.xlsx"
    print("Loading workbook...")
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    sheet = workbook["Posts"]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    
    excel_post_rows = []
    for values in rows:
        record = dict(zip(headers, values, strict=False))
        row_num = _optional_int(record.get("#"))
        if row_num is None:
            continue
        excel_post_rows.append(row_num)
        
    print(f"Total rows in Excel sheet: {len(excel_post_rows)}")
    if excel_post_rows:
        print(f"Excel row numbers range from {min(excel_post_rows)} to {max(excel_post_rows)}")
    
    with connect() as conn:
        db_rows = conn.execute("SELECT source_row_number, status FROM posts WHERE section = 'historical'").fetchall()
        db_imported = {row['source_row_number'] for row in db_rows if row['source_row_number'] is not None}
        db_completed = {row['source_row_number'] for row in db_rows if row['status'] == 'completed' and row['source_row_number'] is not None}
        
    print(f"Total historical posts in DB: {len(db_imported)}")
    print(f"Total completed historical posts in DB: {len(db_completed)}")
    
    not_imported = sorted(list(set(excel_post_rows) - db_imported))
    imported_but_not_completed = sorted(list(db_imported - db_completed))
    
    print(f"Not imported: {len(not_imported)} posts")
    if not_imported:
        print(f"First 10 not imported: {not_imported[:10]}")
    print(f"Imported but not completed/failed: {len(imported_but_not_completed)} posts")
    if imported_but_not_completed:
        print(f"First 10: {imported_but_not_completed[:10]}")

if __name__ == "__main__":
    main()
